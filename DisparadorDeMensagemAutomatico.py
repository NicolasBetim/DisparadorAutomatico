import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import time
import threading
import urllib.parse
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import os
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
import pyperclip

# Variáveis globais
caminho_planilha = None
caminho_arquivo = None
executando = False
thread_envio = None
driver = None
janela_principal = None
mensagem = None
arquivo_anexo_selecionado = None
arquivo_planilha_selecionado = None
status_var = None
progresso = None

# Cores
COR_FUNDO = "#555769"
COR_FRAME = "#CACDFA"
COR_BOTAO = "#8C8FAD"
COR_TEXTO_BOTAO = "#555769"
COR_BOTAO_HOVER = "#71738B"
COR_TEXTO = "#555769"

def on_enter(e):
    e.widget['background'] = COR_BOTAO_HOVER

def on_leave(e):
    e.widget['background'] = COR_BOTAO

def iniciar_driver():
    global driver
    print("Iniciando o driver do Chrome...")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.get("https://web.whatsapp.com/")
        print("Driver iniciado com sucesso. Por favor, escaneie o QR Code do WhatsApp.")
    except Exception as e:
        print(f"Erro ao iniciar o driver: {str(e)}")
        driver = None

def enviar_mensagem_whatsapp(numero, mensagem, arquivo):
    print(f"Tentando enviar mensagem para {numero}: '{mensagem}'")
    global driver
    try:
        numero_formatado = f"+{numero}"
        link = f"https://web.whatsapp.com/send?phone={numero_formatado}"
        driver.get(link)
        
        campo_mensagem = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'))
        )
        
        time.sleep(1)
        
        # Usar pyperclip para lidar com caracteres especiais e emojis
        pyperclip.copy(mensagem)
        
        ActionChains(driver).move_to_element(campo_mensagem).click().perform()
        
        # Usar CTRL+V para colar a mensagem
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
        
        time.sleep(0.5)
        
        if arquivo:
            extensoes_imagem = ['.jpg', '.jpeg', '.png', '.gif']
            _, extensao = os.path.splitext(arquivo.lower())
            
            botao_anexo = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//div[@title="Anexar"]'))
            )
            botao_anexo.click()
            time.sleep(0.5)
            
            if extensao in extensoes_imagem:
                botao_imagem = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]'))
                )
                botao_imagem.send_keys(arquivo)
            else:
                botao_documento = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//input[@accept="*"]'))
                )
                botao_documento.send_keys(arquivo)
            
            time.sleep(2)
        
        for _ in range(3):
            try:
                botao_enviar = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '//span[@data-icon="send"]'))
                )
                botao_enviar.click()
                break
            except (TimeoutException, ElementClickInterceptedException):
                time.sleep(0.5)
        else:
            # Usar ENTER como último recurso
            campo_mensagem.send_keys(Keys.ENTER)
        
        time.sleep(1)
        
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//span[@data-icon="msg-check"]'))
            )
            print(f"Mensagem enviada com sucesso para {numero}")
            return True
        except TimeoutException:
            print(f"Não foi possível confirmar o envio da mensagem para {numero}")
            return False
        
    except Exception as e:
        print(f"Erro ao enviar para {numero}: {str(e)}")
        return False
    
def enviar_mensagens(mensagem, caminho_planilha, caminho_arquivo):
    global executando, driver
    executando = True
    try:
        if not driver:
            iniciar_driver()
        
        workbook = openpyxl.load_workbook(caminho_planilha)
        sheet = workbook.active
        total_contatos = sum(1 for row in sheet.iter_rows(min_row=2) if row[0].value)
        contatos_processados = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not executando:
                break
            numero = str(row[0])
            numero = ''.join(filter(str.isdigit, numero))
            if numero:
                for tentativa in range(2):
                    if enviar_mensagem_whatsapp(numero, mensagem, caminho_arquivo):
                        break
                    elif tentativa < 1:
                        print(f"Tentativa {tentativa + 1} falhou. Tentando novamente em 3 segundos...")
                        time.sleep(3)
                    else:
                        print(f"Falha ao enviar para {numero} após 2 tentativas")
                
                contatos_processados += 1
                progresso_percentual = (contatos_processados / total_contatos) * 100
                atualizar_progresso(progresso_percentual)
                atualizar_status(f"Processado {contatos_processados} de {total_contatos} contatos")
                
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
    finally:
        executando = False
        atualizar_status("Envio de mensagens finalizado.")
        atualizar_progresso(100)

def parar_execucao():
    global executando, driver
    executando = False
    atualizar_status("Parando o envio de mensagens...")
    atualizar_progresso(100)
    if driver:
        driver.quit()
        driver = None

def abrir_arquivo_planilha():
    global caminho_planilha
    caminho_planilha = filedialog.askopenfilename(
         title="Selecione a planilha de contatos",
         filetypes=(("Planilhas Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
    )
    if caminho_planilha:
        arquivo_planilha_selecionado.set(caminho_planilha)
        atualizar_status("Planilha de contatos selecionada.")

def abrir_arquivo_anexo():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
         title="Selecione o arquivo para anexar",
         filetypes=(("Todos os arquivos", "*.*"), ("Documentos", "*.pdf"), ("Imagens", "*.png *.jpg *.jpeg"))
    )
    if caminho_arquivo:
        arquivo_anexo_selecionado.set(caminho_arquivo)
        atualizar_status("Arquivo para anexar selecionado.")

def atualizar_status(mensagem):
    status_var.set(mensagem)
    janela_principal.update_idletasks()

def atualizar_progresso(valor):
    progresso['value'] = valor
    janela_principal.update_idletasks()

def iniciar_progresso():
    progresso['value'] = 0

def parar_progresso():
    progresso['value'] = 100

def enviar_mensagens_interface():
    global thread_envio
    mensagem_texto = mensagem.get("1.0", tk.END).strip()
    if caminho_planilha and mensagem_texto:
        iniciar_progresso()
        atualizar_status("Enviando mensagens...")
        thread_envio = threading.Thread(target=enviar_mensagens, args=(mensagem_texto, caminho_planilha, caminho_arquivo))
        thread_envio.start()
    else:
        messagebox.showwarning("Aviso", "Por favor, selecione a planilha de contatos e insira uma mensagem.")

def criar_interface():
    global janela_principal, mensagem, arquivo_anexo_selecionado, arquivo_planilha_selecionado, status_var, progresso

    janela_principal = tk.Tk()
    janela_principal.title("Automação de Mensagens WhatsApp")
    janela_principal.configure(bg=COR_FUNDO)

    main_frame = ttk.Frame(janela_principal, padding="10", style='TFrame')
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    style = ttk.Style()
    style.configure('TFrame', background=COR_FRAME)
    style.configure('TButton', background=COR_BOTAO, foreground=COR_TEXTO_BOTAO)
    style.configure('TLabel', background=COR_FRAME, foreground=COR_TEXTO)

    arquivo_planilha_selecionado = tk.StringVar()
    arquivo_anexo_selecionado = tk.StringVar()
    status_var = tk.StringVar()

    ttk.Label(main_frame, text="Planilha de Contatos:").grid(row=0, column=0, sticky=tk.W, pady=5)
    ttk.Entry(main_frame, textvariable=arquivo_planilha_selecionado, width=40).grid(row=0, column=1, pady=5)
    ttk.Button(main_frame, text="Selecionar", command=abrir_arquivo_planilha).grid(row=0, column=2, pady=5)

    ttk.Label(main_frame, text="Arquivo Anexo:").grid(row=1, column=0, sticky=tk.W, pady=5)
    ttk.Entry(main_frame, textvariable=arquivo_anexo_selecionado, width=40).grid(row=1, column=1, pady=5)
    ttk.Button(main_frame, text="Selecionar", command=abrir_arquivo_anexo).grid(row=1, column=2, pady=5)

    ttk.Label(main_frame, text="Mensagem:").grid(row=2, column=0, sticky=tk.W, pady=5)
    mensagem = tk.Text(main_frame, height=5, width=40)
    mensagem.grid(row=2, column=1, columnspan=2, pady=5)

    ttk.Button(main_frame, text="Iniciar Envio", command=enviar_mensagens_interface).grid(row=3, column=1, pady=10)
    ttk.Button(main_frame, text="Parar Envio", command=parar_execucao).grid(row=3, column=2, pady=10)

    ttk.Label(main_frame, textvariable=status_var).grid(row=4, column=0, columnspan=3, pady=5)
    ttk.Label(main_frame, text="Criado por Nicolas Betim.", font=("arial",8,"italic")).grid(row=6, column=1)

    progresso = ttk.Progressbar(main_frame, mode='determinate', length=300)
    progresso.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)

    janela_principal.columnconfigure(0, weight=1)
    janela_principal.rowconfigure(0, weight=1)
    main_frame.columnconfigure(1, weight=1)
    for i in range(6):
        main_frame.rowconfigure(i, weight=1)

def main():
    print("Iniciando o programa...")
    iniciar_driver()
    criar_interface()
    janela_principal.mainloop()

if __name__ == "__main__":
    main()
    